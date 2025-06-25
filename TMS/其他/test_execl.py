from openpyxl import load_workbook

wb = load_workbook(u"../其他/test12.xlsx")
sheetnames = wb.sheetnames # 获得表单名字
x = "TESTIT20230214001-01"
sheet = wb[sheetnames[0]]
for i in range(2,25225):#ws.max_row
    if i % 100 == 0:
        x="TESTIT20230214001-"+str(i)
    sheet['A'+str(i)] = x
wb.save(u"../其他/test12.xlsx")
