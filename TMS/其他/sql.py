import openpyxl



wb = openpyxl.load_workbook('../其他/phone.xlsx')#读取execl订单后入库
ws = wb.active
lie = ws.max_column
hang = ws.max_row

name = "test"

fields = "('編號','條碼')"

for x in range(1,ws.max_row+1):
    data = []
    for i in range(1,ws.max_column+1):
        cols = ws.cell(x, i).value
        data.append(cols)
    sql = "INSERT INTO " + name + fields + " VALUES ("
    for b in range(0,len(data)):
        if b == len(data)-1:
            sql = sql +"'"+ str(data[b]) + "');"
        else:
            sql = sql +"'" + str(data[b]) + "',"
    print(sql)





# for i in range(1,ws.max_column):
#     data = []
#     for x in range(1,ws.max_row):
#         cols = ws.cell(1, x).value
#         data.append(cols)
#     print(data)
#     for b in range(0,len(data)):
#         print("INSERT INTO act_status_pull (A,B) VALUES ("+str(data[b])+")")
#     data = []
#     print(data)
#     A=str(ws['A'+str(i)].value)
#     B=str(ws['B'+str(i)].value)
    # print("INSERT INTO act_status_pull (A,B) VALUES (" + "'"+ A + "',"+B+")")