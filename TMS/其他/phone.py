import json
import random

import openpyxl
import requests
import time

import json

import requests


def login():
    url = "http://47.119.120.7:8000/pos-web/token/get"#测试
    #url = "http://120.78.66.231:8000/pos-web/token/get" #生产
    payload={
                "username": "999666_KERRYCN",
                "password": "b05b41732aac4fa491723669c35f10d3"
            }
    headers = {
      'Content-Type': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
    return json.loads(response.text)['body']['token']


def file_create_order(token):
    header = {
        'Content-Type':'application/json',
        "Authorization":"Bearer"+" "+token
        }
    with open("../file_data/order_data.txt", 'r',encoding= 'utf-8') as f:
        param2 = json.loads(f.read())#转换成字典
        f.close()
    reference_number="TESTBACK"+str(random.randint(1,9999999999))
    # reference_number = "LPO212211221"
    param2['package']['reference_number']=reference_number
    url = "http://47.119.120.7:8000/pos-web/shipment/create"
    time_start = time.time()
    response = requests.post(url ,data=json.dumps(param2), headers = header)
    time_end = time.time()
    print('下单耗时：', round(time_end - time_start, 2), 's')
    if response.status_code ==201:
        print(response.text)
        # print(json.loads(response.text)["data"]["label_url"])
        print(json.loads(response.text)["data"]["tracking_number"])
        # success_length = open('../request_data/success_length.txt', 'ab')
        # success_length.write(((json.loads(response.text)["data"]["tracking_number"])+"\n").encode('utf-8'))
        return json.loads(response.text)["data"]["tracking_number"]
        # success_length.close()
    else:
        print(reference_number)
        print(response.text)
        return "失败"

def open_excel():#读取ececl数据
    token = login()
    wb = openpyxl.load_workbook('../其他/phone.xlsx')#TH泰国参数
    ws = wb.active
    data = []
    for i in range(1,104):#ws.max_row
        x=str(ws['A'+str(i)].value)
        data.append(x)
        header = {
            'Content-Type':'application/json',
            "Authorization":"Bearer"+" "+token
            }
        with open("../file_data/order_data.txt", 'r',encoding= 'utf-8') as f:
            param2 = json.loads(f.read())#转换成字典
            f.close()
        reference_number="TESTBACK"+str(random.randint(1,9999999999))
        param2['package']['reference_number']=reference_number
        param2['receiver']['phone'] = x
        url = "http://47.119.120.7:8000/pos-web/shipment/create"
        response = requests.post(url ,data=json.dumps(param2), headers = header)
        if response.status_code ==201:
            print("seccess")
            # print(json.loads(response.text)["data"]["tracking_number"])
        else:
            print(reference_number + "失败："+x)
        wb.close()

