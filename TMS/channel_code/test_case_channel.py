import unittest
from time import sleep

import openpyxl

from TMS.channel_code.public.channel import add_channel, select_hubInId
from TMS.channel_code.public.hubout import add_hubout, select_hubOutId
from TMS.channel_code.public.sort_code import add_sortcode


class MyTestCase(unittest.TestCase):
    def test_something(self):
        self.assertEqual(True, False)


    def test_case_order_create(self):
        for i in range(1):
            # wb = openpyxl.load_workbook('channel.xlsx')#读取execl订单后入库
            # ws = wb.active
            # name = str(ws['A'+str(i)].value)
            # even_mess = str(ws['B'+str(i)].value)
            name = "VNHCMRT-TEST2"
            country = "TH"
            # sortCode = "TH-15"
            # name = str(ws['A'+str(i)].value)
            # country = str(ws['B'+str(i)].value)
            sortCode = name
            add_channel(name,country)
            add_hubout(name = name,country = country)
            sleep(5)
            hubInId = select_hubInId(name)
            hubOutId =select_hubOutId(name=name)
            add_sortcode(name = name,country = country,sortCode =sortCode,hubOutId = hubOutId,hubInId = hubInId)


if __name__ == '__main__':
    unittest.main()
