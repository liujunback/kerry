import json

import openpyxl

from case.public.pdf import pdf


def create_pdf():
        wb = openpyxl.load_workbook('karcel_dev.xlsx')
        ws = wb.active
        data = []
        for i in range(2,501):
            tracking = str(ws['A'+str(i)].value)
            url = str(ws['B'+str(i)].value)
            print(tracking+"  "+url)
            pdf(url=url,a=tracking)



create_pdf()