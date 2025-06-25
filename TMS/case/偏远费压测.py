import json
import random

import openpyxl
import requests
from locust import HttpUser,TaskSet,task


class Test(TaskSet):#偏远费
    def on_start(self):

        r2 = requests.post("http://120.24.31.239:20000/tms-saas-web/user/login?userNo=zmuser&password=123456&companyNo=&domain=")
        self.token = json.loads(r2.text)["body"]["token"]
        wb = openpyxl.load_workbook('../file_data/偏远费标准导入模板测试数据.xlsx')#TH泰国参数
        ws = wb.active
        self.TH_zip = []
        self.TH_fee = []
        for i in range(4,28):
            x=str(ws['G'+str(i)].value)
            fee=str(ws['K'+str(i)].value)
            self.TH_zip.append(x)
            self.TH_fee.append(fee)
        self.SG_zip = []
        self.SG_fee = []
        for i in range(47,162):
            x=str(ws['G'+str(i)].value)
            fee=str(ws['K'+str(i)].value)
            self.SG_zip.append(x)
            self.SG_fee.append(fee)
        self.VN_city = []
        self.VN_district = []
        for i in range(163,9063):
            city=str(ws['E'+str(i)].value)
            district=str(ws['F'+str(i)].value)
            self.VN_city.append(city)
            self.VN_district.append(district)

    @task(2)
    def Free_TH(self):#偏远费    下单
        zip = random.randint(0,23)
        payload={
            "postalType":"TH",
            "countryCode":"TH",
            "district":"",
            "city":"",
            "addressKey":"",
            "zip":self.TH_zip[zip],
            "weig":"1",
            "token" : self.token
        }
        headers = {
          'Content-Type': 'application/x-www-form-urlencoded;charset=UTF-8'
        }
        with self.client.post('/tms-saas-web/bms/lineremotefee/find', data=payload ,headers=headers ,name = "偏远费TH", catch_response = True) as response:
            if self.TH_fee[zip] in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(payload).encode('utf-8'))
                failder_data.write(",")
                failder_data.close()

    @task(2)
    def Free_SG(self):#不走末端    下单
        zip = random.randint(0,110)
        payload={
            "postalType":"SG",
            "countryCode":"SG",
            "district":"",
            "city":"",
            "addressKey":"",
            "zip":self.SG_zip[zip],
            "weig":"1",
            "token" : self.token
        }
        headers = {
          'Content-Type': 'application/x-www-form-urlencoded;charset=UTF-8'
        }
        with self.client.post('/tms-saas-web/bms/lineremotefee/find', data=payload ,headers=headers ,name = "偏远费SG", catch_response = True) as response:
            if self.SG_fee[zip] in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(payload).encode('utf-8'))
                failder_data.write(",")
                failder_data.close()

    @task(2)
    def Free_VN(self):#不走末端    下单
        num = random.randint(0,8000)
        payload={
            "postalType":"VN",
            "countryCode":"VN",
            "district":self.VN_district[num],
            "city":self.VN_city[num],
            "addressKey":"",
            "zip":"",
            "weig":"1",
            "token" : self.token
        }
        headers = {
          'Content-Type': 'application/x-www-form-urlencoded;charset=UTF-8'
        }
        with self.client.post('/tms-saas-web/bms/lineremotefee/find', data=payload ,headers=headers ,name = "偏远费VN", catch_response = True) as response:
            if "5" in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(payload).encode('utf-8'))
                failder_data.write(",")
                failder_data.close()







class websitUser(HttpUser):
    tasks = [Test]
    host = "http://120.24.31.239:20000"
    min_wait = 1000  # 单位为毫秒
    max_wait = 2000  # 单位为毫秒
