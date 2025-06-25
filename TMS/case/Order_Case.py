import ast
import json
import os
import random
import urllib

import openpyxl
import requests

from locust import HttpUser,TaskSet,task


class Test(TaskSet):

    def on_start(self):
        #读取表格数据
        wb = openpyxl.load_workbook('../file_data/TH_data.xlsx')#TH泰国参数
        ws = wb.active
        self.data = []
        for i in range(2,2001):
             x=str(ws['A'+str(i)].value)
             self.data.append(json.loads(x))

        wb_hk = openpyxl.load_workbook('../file_data/HK_data.xlsx')#HK香港参数
        HK = wb_hk.active
        self.hk_data = []
        for i in range(2,300):
             x=str(HK['A'+str(i)].value)
             self.hk_data.append(json.loads(x))

        wb_SG = openpyxl.load_workbook('../file_data/SG_data.xlsx')#SG新加坡参数
        SG = wb_SG.active
        self.sg_data = []
        for i in range(2,200):
             x=str(SG['A'+str(i)].value)
             self.sg_data.append(json.loads(x))

        wb_MY = openpyxl.load_workbook('../file_data/MY_data.xlsx')#MY马来西亚参数
        MY = wb_MY.active
        self.my_data = []
        for i in range(2,200):
             x=str(MY['A'+str(i)].value)
             self.my_data.append(json.loads(x))

        #登录获取token
        #url = "http://47.119.120.7:8000/pos-web/token/get"
        url = "http://120.78.66.231:8000/pos-web/token/get"
        payload={
                        "username": "01_KERRYCN",
                        "password": "c24e24b2d2e541f19da6acb5f1af9298"
                    }
        headers = {
          'Content-Type': 'application/json'
        }
        response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
        self.token = json.loads(response.text)['body']['token']
        print(self.token)

    @task(2)
    def create_order_TH(self):#下单
        # 定义请求头
        header = {
            'Content-Type':'application/json',
            "Authorization":"Bearer"+" "+self.token
            }

        #data参数
        datas = self.data[random.randint(1,100)]
        datas['sale_platform'] = ''
        #channel_code = ["CACNTHG00","CACNTHG01","LACNTHG00"]
        datas['service']['channel_code'] = "CACNTHG01"
        datas['receiver']['country_code'] = "TH"
        datas['package']['payment_method'] = "PP"
        datas['receiver']['post_code'] = "24000"
        ref_num = datas['receiver']['country_code'] + "051501" + str(random.randint(1,99999999999999))
        datas['package']['reference_number'] = ref_num
        datas['items'][0]['description_origin_language'] = "CN"
        #下单
        with self.client.post('/pos-web/shipment/create', data = json.dumps(datas), headers = header, name = "泰国下单", catch_response = True) as response:
            if ref_num in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_response = open('../request_data/failed_response.txt', 'ab')
                failder_response.write(str(response.text).encode('utf-8'))
                failder_response.close()
                datas = json.dumps(datas)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(datas+"\n").encode('utf-8'))
                failder_data.close()

    @task(2)
    def create_order_HK(self):#下单
        # 定义请求头
        headers = {
          'Content-Type': 'application/json'
        }
        #data参数
        datas = self.hk_data[random.randint(1,100)]
        datas['sale_platform'] = ''
        #channel_code=["CTCNHK001","CTCNHK000"]
        datas['service']['channel_code'] = "CTCNHK000"
        datas['receiver']['country_code'] = "HK"
        datas['package']['payment_method'] = "PP"
        ref_num = datas['receiver']['country_code'] + "051501" + str(random.randint(1,99999999999999))
        datas['package']['reference_number'] = ref_num
        datas['items'][0]['description_origin_language']="时钟"

        #下单
        with self.client.post('/pos-web/shipment/create', data = json.dumps(datas), headers = headers, name = "香港下单", catch_response = True) as response:
            if ref_num in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_response = open('../request_data/failed_response.txt', 'ab')
                failder_response.write(str(response.text).encode('utf-8'))
                failder_response.close()
                datas = json.dumps(datas)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(datas+"\n").encode('utf-8'))
                failder_data.close()

    @task(2)
    def create_order_SG(self):#下单
        # 定义请求头
        header = {
            'Content-Type':'application/json',
            "Authorization":"Bearer"+" "+self.token
            }

        #data参数
        datas = self.sg_data[random.randint(1,100)]
        datas['sale_platform'] = ''
        #channel_code=["CACNSGG00","CACNSGG01"]
        datas['service']['channel_code'] = "CACNSGG00"
        datas['receiver']['country_code'] = "SG"
        datas['package']['payment_method'] = "PP"
        ref_num = datas['receiver']['country_code'] + "051501" + str(random.randint(1,99999999999999))
        datas['package']['reference_number'] = ref_num
        datas['items'][0]['description_origin_language']="CN"
        #datas['items'][0]['unit_price'] = 666
        #下单
        with self.client.post('/pos-web/shipment/create', data = json.dumps(datas), headers = header, name = "新加坡下单", catch_response = True) as response:
            if ref_num in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_response = open('../request_data/failed_response.txt', 'ab')
                failder_response.write(str(response.text).encode('utf-8'))
                failder_response.close()
                datas = json.dumps(datas)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(datas+"\n").encode('utf-8'))
                failder_data.close()

    @task(2)
    def create_order_MY(self):#下单
        # 定义请求头
        header = {
            'Content-Type':'application/json',
            "Authorization":"Bearer"+" "+self.token
            }

        #data参数
        datas = self.my_data[random.randint(1,100)]
        datas['sale_platform'] = ''
        datas['service']['channel_code'] ="CACNMYG00"
        datas['receiver']['country_code'] = "MY"
        datas['receiver']['name']='back'
        datas['package']['payment_method'] = "PP"
        ref_num = datas['receiver']['country_code'] + "050801" + str(random.randint(1,99999999999999))
        datas['package']['reference_number'] = ref_num
        datas['items'][0]['description_origin_language']="CN"
        #datas['items'][0]['unit_price'] = 666
        #下单
        with self.client.post('/pos-web/shipment/create', data = json.dumps(datas), headers = header, name = "马来西亚下单", catch_response = True) as response:
            if ref_num in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_response = open('../request_data/failed_response.txt', 'ab')
                failder_response.write(str(response.text).encode('utf-8'))
                failder_response.close()
                datas = json.dumps(datas)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(datas+"\n").encode('utf-8'))
                failder_data.close()

    @task(1)
    def order_status(self):#查货态
        # 定义请求头
        header = {
            'Content-Type':'application/json',
            "Authorization":"Bearer"+" "+self.token
            }
        tracking_number = "KENLNT00109873"
        with self.client.get('/pos-web/shipment/status?tracking_number='+tracking_number, headers = header, name = "查货态", catch_response = True) as response:
            #print(response.text)
            if "success" in response.text:
                response.success()
            else:
                response.failure("")
                f = open('failed_response.txt', 'a')
                # os.linesep
                f.write(str(response.text))




class websitUser(HttpUser):
    tasks = [Test]
    host = "0"
    min_wait = 1000  # 单位为毫秒
    max_wait = 2000  # 单位为毫秒