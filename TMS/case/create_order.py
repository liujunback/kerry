import json
import threading
import unittest

import time

import  warnings

import pandas as pd

from TMS.public.spider import spider

warnings.filterwarnings('ignore')
import datetime

import openpyxl

from TMS.ICBU.package_scan import package_scan
from TMS.ICBU.status import status
from TMS.public.Mawb import create, scan_box, close_mawb
from TMS.public.TwmsLogin import Twms_inbound, Twms_put

from TMS.public.check_weight import check_weight
from TMS.public.close_box_scan import close_Box_Scan,close_Box
from TMS.public.create_order import file_create_order
from TMS.public.Login import login
from TMS.public.Inbaound import inbound
from TMS.public.delivery import delivery
from TMS.public.shipment import shipment_add, shipment_num_ids, shipment_scan, shipment_close



class MyTestCase(unittest.TestCase):

    def setUp(self):
        pass
    def tearDown(self):
        pass
    # @unittest.skip("")#包含WMS的退件入库和上架接口CTCNTH000123123
    def test_case_order_create(self):
        fail = 0
        x = 2
        trak=[]
        box_list=[]
        box_num = 0
        token = login()#be08f7c5-a2c1-4823-a869-c629c2efdc5e
        print(token)
        shipment_num=0        # token = "aacc2b37-d5f4-4f3e-9a26-12cae1320e7a"
        if x>=1:
            for i in range(1):
                # tracking_num = file_create_order(token)
                tracking_num = "KECTH92000135"
                if tracking_num == "失败":
                    fail=fail+1
                # tracking_num = "CTCNTH000"
                else:
                    trak.append(tracking_num)
            print(fail)
            if x>=2:
                for i in range(len(trak)):
                    if i == 0:
                        # time.sleep(30)
                        # spider(trak[i])
                        # package_scan(trak[i])
                        inbound(trak[i])
                        time.sleep(1)
                        box_num = close_Box_Scan(trak[i])
                        box_list.append(box_num)
                    else:
                        # spider(trak[i])
                        inbound(trak[i])
                        close_Box_Scan(trak[i],box_num)
                close_Box(box_num,trak)
                time.sleep(10)
                check_weight(box_num,trak)
                shipment_num = shipment_add()
                mawb_data = create()
                time.sleep(30)
                shipmentbatchId = shipment_scan(box_num,shipment_num)
                shipment_close(shipmentbatchId,shipment_num)
                # #
                time.sleep(10)
                scan_box(box_num,mawb_data["mawb"],mawb_data["id"])
                close_mawb(mawb_data["mawb"],mawb_data["id"])
                # time.sleep(10)
                # status(tracking_num,"OC","航班起飞")
                # time.sleep(10)
                # status(tracking_num,"OQ","进口清关完成")
       # for i in range(len(trak)):
       #      text = Twms_inbound(trak[i])
       #      Twms_put(trak[i])
                # status(tracking_num,"LS","丢失")
                # status(tracking_num,"OC","航班起飞")
                # status(tracking_num,"OK","快件已签收")
        # shipment_close(shipmentbatchId,shipment_num)
        # tracking_num="KENLNT00184371"

        # status(tracking_num,"RT","退回PDD CN仓库")
        # status(tracking_num,"ON","未放行")
        # status(tracking_num,"OE","放行")
        # status(tracking_num,"RJ","收件⼈拒绝签收")
        # time.sleep(3)
        # for i in range(len(trak)):
        # #     # status(trak[i],"DT",tms_token,properties,"干线交接")#"EH", "EN", "OG", "ON", "SP3F", "SP2F", "XH", "LS", "DM"
        #     time.sleep(1)
        #     status(trak[i],"ES","出口报关开始")#CAINIAO_GLOBAL_CCEX_START_CALLBACK
        # #     # time.sleep(1)
        # #     # status(trak[i],"EH","出口清关查件")
        # #     # time.sleep(1)
        # #     # status(trak[i],"EN","违禁品")
        #     time.sleep(1)
        #     status(trak[i],"FX","出口清关完成")
        #     time.sleep(1)
        #     status(trak[i],"OC","航班起飞")#CAINIAO_GLOBAL_LINEHAUL_DEPARTURE_CALLBACK
        #     time.sleep(1)
        #     status(trak[i],"OF","航班抵达")
        #     time.sleep(1)
        #     status(trak[i],"OR","入口清关收到货")
        #     time.sleep(1)
        #     status(trak[i],"OS","进口清关开始")
        # #     # time.sleep(1)
        # #     # status(trak[i],"OG","进口清关查件")
        # #     # time.sleep(1)
        # #     # status(trak[i],"ON","进口清关销毁")
        #     time.sleep(1)
        #     status(trak[i],"OQ","进口清关完成")
        # #     # time.sleep(1)
        # #     # status(trak[i],"HL","交货到末公里")
        #     time.sleep(1)
        #     status(trak[i],"ZY","包裹到达分拣中心")
        #     time.sleep(1)
        #     status(trak[i],"SP","站点发出")
        # #     # time.sleep(1)
        # #     # status(trak[i],"SP1F","首次派送失败")
        # #     # time.sleep(1)
        # #     # status(trak[i],"SP2","站点发出")
        # #     # time.sleep(1)
        # #     # status(trak[i],"SP2F","二次派送失败")
        # #     # time.sleep(1)
        # #     # status(trak[i],"SP3","三次派送")
        # #     # status(trak[i],"SP3F","三次派送失败")
        #     time.sleep(1)
        #     status(trak[i],"OK","用户签收")
            # time.sleep(1)
            # status(trak[i],"ZU","退件上架")
            # # status(trak[i],"RJ","拒收")
            # # status(trak[i],"RN","lastmile_eturn")
            # time.sleep(1)
            # status(trak[i],"XH","销毁")
            # status(trak[i],"LS","丢失")
            # time.sleep(1)
            # status(trak[i],"TH",tms_token,properties,"退回")
            # status(trak[i],"DM",tms_token,properties,"DAMAGE")










    @unittest.skip("")#包含WMS的退件入库和上架接口
    def test_case_order_create1(self):#下单
        box_list = 100
        token = login()
        #tracking_number = creat_order(token)#ececl下单
        print(token)
        trackingnumber_list = []
        for list in range(box_list):
            wb = openpyxl.load_workbook('../main/test.xlsx')#读取execl订单后入库
            ws = wb.active
            tr=str(ws['A'+str(list)].value)
            tracking_num = tr
            # tracking_num = file_create_order(token)#读取文件下单
            #tracking_num = "KENLNT00177492"
            trackingnumber_list.append(tracking_num)
        # #delivery(tracking_num)#收货
        box_num=""
        for tra in range(box_list):
            text=""
            if tra == 0:
                time.sleep(25)
                inbound(trackingnumber_list[tra])#入库
                from TMS.public.TwmsLogin import Twms_inbound
                text = Twms_inbound(trackingnumber_list[tra])
                if "successfully" in text:
                        failder_data = open('../request_data/failed_data.txt', 'ab')
                        failder_data.write((str(trackingnumber_list[tra])+"\n").encode('utf-8'))
                        failder_data.close()
            else:
                inbound(trackingnumber_list[tra])#入库
                text = Twms_inbound(trackingnumber_list[tra])
                # if "successfully" in text:
                #         failder_data = open('../request_data/failed_data.txt', 'ab')
                #         failder_data.write((str(trackingnumber_list[tra])+"\n").encode('utf-8'))
                #         failder_data.close()
                #     box_num = close_Box_Scan(trackingnumber_list[tra])#关联箱子
                # else:
                #     inbound(trackingnumber_list[tra])
                #     close_Box_Scan(trackingnumber_list[tra],box_num)
            close_Box(box_num,trackingnumber_list)#关箱
            check_weight(box_num,trackingnumber_list)#核重
            shipment_num = shipment_add()
            time.sleep(15)
            shipmentbatchId = shipment_scan(box_num,shipment_num)
            shipment_close(shipmentbatchId,shipment_num)

    # @unittest.skip("")
    def test_case_order_create12(self):
        num = 100
        trakings=[]
        box_num_list = []
        shipment_num = shipment_add()
        mawb = create()
        for i in range(num):
            tracking_num = file_create_order(login())
            # wb = openpyxl.load_workbook('../case/test.xlsx', read_only=True, data_only=True)#读取execl订单后入库
            # ws = wb.active
            # tr=str(ws['A'+str(i)].value)
        # df = pd.read_excel('../case/test.xlsx', engine='openpyxl',nrows=100000)
        # for index, row in df.iterrows():
        #     tr = row['number']
        #     print(str(index) + "行数" + row['number'])
        #     tracking_num = tr
            trakings.append(tracking_num)
        track_list = [trakings[i:i+10] for i in range(0,len(trakings),10)]
        time.sleep(30)
        for x in track_list:
            one = 0
            box_num = 0
            cou= len(x)
            for track in x:
                print(track)
                one = int(x.index(track))
                if one == 0:
                    if one == (cou-1):
                        time.sleep(1)
                        inbound(track)
                        box_num = close_Box_Scan(tracking_number=track)
                        close_Box(box_num,x)
                        check_weight(box_num,x)
                        box_num_list.append(box_num)
                    else:
                        # time.sleep(30)
                        inbound(track)
                        box_num = close_Box_Scan(tracking_number=track)
                elif one == (cou-1):
                    inbound(track)
                    close_Box_Scan(track,box_num)
                    close_Box(box_num,x)
                    check_weight(box_num,x)
                    box_num_list.append(box_num)
                else:
                    inbound(track)
                    close_Box_Scan(track,box_num)
        shipmentbatchId = 0
        time.sleep(10)
        for i in range(len(box_num_list)):
            shipmentbatchId = shipment_scan(box_num_list[i],shipment_num)
        shipment_close(shipmentbatchId,shipment_num)
        for i in range(len(box_num_list)):
            scan_box(box_num_list[i],mawb["mawb"],mawb["id"])
        close_mawb(mawb["mawb"],mawb["id"])




if __name__ == '__main__':
    unittest.main()