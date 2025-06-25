import json

import openpyxl
import requests

def open_excel():#读取ececl数据
    wb = openpyxl.load_workbook('../JD/JD_data.xlsx')#TH泰国参数
    ws = wb.active
    data = []
    for i in range(20,21):#ws.max_row
        print(i)
        x=str(ws['A'+str(i)].value)
        if x !="NULL":
            data.append(json.loads(x))
    return data


def JD_create():
    datas = open_excel()
    req = {
                "response":{
                    "content":{
                        "data":datas[0],
                        "exception":"",
                        "status":"SUCCESS"
                    },
                    "code":0
                }
            }
    print(json.dumps(req))

