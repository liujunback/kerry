import openpyxl
import requests
import time

from TMS.ICBU.status import status
from TMS.public.Mawb import create, scan_box, close_mawb
from TMS.public.check_weight import check_weight
from TMS.public.close_box_scan import close_Box_Scan,close_Box
from TMS.public.Login import login
from TMS.public.Inbaound import inbound
from TMS.public.shipment import shipment_add, shipment_scan, shipment_close


def open_excel():#读取ececl数据
    wb = openpyxl.load_workbook('../JD/jd_orders.xlsx')
    ws = wb.active
    data = []
    for i in range(10101,10102):#ws.max_row
        x=str(ws['A'+str(i)].value)
        if x !="NULL":
            data.append(x)
    return data



def create_JD_order(order_type,ddl):

    url = "http://120.24.31.239:20000/tms-saas-web/api/jdwl?"
    data = "msg_id=lsb-g3%2Ctask%2C548599634%2Cmagellan-channel-pro%2C1%2C0&msg_name=jingdong_magellan_pushdown_channel&msg_payload=%7B%22orderType%22%3A%22"+order_type+"%22%2C%22orderSource%22%3A%22JOYBUY%E4%B8%8B%E5%8D%95%22%2C%22orderNumber%22%3A%22"+ddl+"%22%2C%22orderCreateTime%22%3A%22Jul+30%2C+2021+2%3A22%3A03+PM%22%2C%22ddl%22%3A%22"+ddl+"%22%7D&msg_pin=jd_51676128fefa1"

    payload={}
    headers = {}
    url = url+data
    response = requests.request("POST",url , headers=headers, data=payload)
    print(url)
    print(response.text)

order_type = '1'
ddl = open_excel()
for i in ddl:
    create_JD_order(order_type,i)
# create_JD_order(order_type,ddl)


#
tracking_number= "KECTH91020671"#补录货态
# trak=[]
# box_num = "TH1121110000007M"
# token = login()
# trak.append(tracking_number)
# inbound(trak[0])
# box_num = close_Box_Scan(trak[0])
# close_Box(box_num,trak)
# check_weight(box_num,trak)
# shipment_num = shipment_add()
# time.sleep(15)
# shipmentbatchId = shipment_scan(box_num,shipment_num)
# shipment_close(shipmentbatchId,shipment_num)
# mawb_data = create()
# time.sleep(2)
# scan_box(box_num,mawb_data["mawb"],mawb_data["id"])
# close_mawb(mawb_data["mawb"],mawb_data["id"])
# status(tracking_number,"OC","航班起飞")
# time.sleep(60)
# status(tracking_number,"OF","航班抵达")
# time.sleep(60)
# status(tracking_number,"OQ","⽬的地清关完成")
# time.sleep(60)
# status(tracking_number,"HL","到达转运中心")
# time.sleep(60)
# status(tracking_number,"IT","离开转运中⼼")
# time.sleep(60)
# status(tracking_number,"SP","安排投递")
# time.sleep(60)
# status(tracking_number,"RJ","收件⼈拒绝签收")
# time.sleep(60)
# status(tracking_number,"OK","快件已签收")