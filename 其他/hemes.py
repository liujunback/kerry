import datetime
from time import sleep

import openpyxl

import json

for i in range(2,492):
    wb = openpyxl.load_workbook('../其他/hermes2.xlsx')#读取execl订单后入库
    ws = wb.active
    event = str(ws['A'+str(i)].value)
    even_mess = str(ws['B'+str(i)].value)
    import requests
    url = "http://stg.spider.tec-api.com:38005/package/tracking/hermesborderguru"
    data = datetime.datetime.now().strftime('%H:%M:%S')
    # print(data)
    payload ={
                "barcode":"H1234560013568801031",
                "orderId":"H1234560013568801031",
                "eventType":event,
                "eventMessage":even_mess,
                "trackingTimestamp":"2023-09-13T"+data+".9905706"
            }
    headers = {
      'Authorization': 'Bearer test123token',
      'Content-Type': 'application/json'
    }
    sleep(1)
    response = requests.request("POST", url, headers=headers, data= json.dumps(payload))

    print(response.text)
