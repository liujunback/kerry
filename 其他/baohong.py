import json

import openpyxl
import requests


wb = openpyxl.load_workbook('../其他/baohogn.xlsx')#读取execl订单后入库


for i in range (2,456):
    ws = wb.active
    tr=str(ws['C'+str(i)].value)
    request = tr

    url = "http://exoms.globex.cn/default/product-soap/web-service"
    payload = request.encode('utf-8')
    sku = str(ws['A'+str(i)].value)
    headers = {
      'Content-Type': 'application/xml'
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    if sku in response.text:
        print(sku)
    else:
        print(payload)
        print(response.text)

