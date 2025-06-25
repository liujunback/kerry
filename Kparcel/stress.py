import json
import random
import urllib

import openpyxl
import requests
from locust import HttpUser,TaskSet,task


class Test(TaskSet):#偏远费
    def on_start(self):
        self.wish_data = []
        wb = openpyxl.load_workbook('../TMS/Wish/test.xlsx')
        ws = wb.active
        for i in range(100):#ws.max_row
            x=str(ws['A'+str(4)].value)
            self.wish_data.append(json.loads(x))
        self.cainiao_data = []
        wb = openpyxl.load_workbook('../TMS/cainiao/cainiao_cex_draft_orders (1).xlsx')
        ws = wb.active
        for i in range(2,100):
             x=json.loads(str(ws['D'+str(i)].value))
             self.cainiao_data.append(x)
        self.Order_data=[]
        for i in range(2,100):#
            with open("../TMS/file_data/order_data.txt", 'r',encoding= 'utf-8') as f:
                param2 = json.loads(f.read())#转换成字典
                self.Order_data.append(param2)
                f.close()
        url = "http://120.79.147.221:8001/pos-web/token/get"#Kparcel接口

        payload={
            "username":"999666_K-PARCEL",
            "password": "c109d37f62934d85863b53fe0d0ac494"
        }
        headers = {
          'Content-Type': 'application/json'
        }
        response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
        self.token = json.loads(response.text)['body']['token']


    @task(2)
    def Wish_Order(self):#Wish_Create
        data1 = self.wish_data[random.randint(0,len(self.wish_data)-2)]
        data1["tracking_id"] ="WOSP0"+str(random.randint(1000000,99999999))
        data1["api_key"] = "r7ozJZmjvQHONMQZ9FIdCgjOxL02lWkaC55AGEQdJnPdlAqYhXbWRhPM6VUA"
        payload=json.dumps(data1)
        headers = {
              'Content-Type': 'application/json'
            }
        with self.client.post('/tms-saas-web/wish/order/create', data = payload, headers = headers, name = "Wish下单", catch_response = True ) as response:
            if json.loads(response.text)["code"] == 0:
                response.success()

    @task(2)
    def CaiNiao_Order(self):#菜鸟订单
        logistics_interface = self.cainiao_data[random.randint(0,len(self.cainiao_data)-2)]
        tracking_number =  "621000000000" + str(random.randint(100000000000000,900000000000000))
        logistics_interface["trackingNumber"] = tracking_number
        logisticsOrderCode = "LP004221" + str(random.randint(100000,999999999))
        logistics_interface["logisticsOrderCode"] = logisticsOrderCode
        payload={'data_digest': 'fOzYH+3L0d7LZya9mURGsQ==',
                'msg_type': '12',
                'logistics_interface': json.dumps(logistics_interface) ,
                'partner_code': '123',
                'from_code': '123',
                'msg_id': '123'}
        files=[
        ]
        headers = {
          'Content-Type': 'application/x-www-form-urlencoded'
        }
        with self.client.post('/tms-saas-web/kparcel/cainiao/order?channel_code=CACNPTKPE', data = urllib.parse.urlencode(payload),files=files , headers = headers, name = "菜鸟下单", catch_response = True ) as response:
            if json.loads(response.text)["success"] == True:
                response.success()
    @task(2)
    def API_Order(self):
        headers = {
                    'Content-Type':'application/json',
                    "Authorization":"Bearer"+" "+self.token
                    }
        param = self.Order_data[random.randint(0,len(self.Order_data)-2)]
        reference_number="THBACK"+str(random.randint(1,99999999999999))
        param['package']['reference_number']=reference_number
        with self.client.post(url = 'http://120.79.147.221:8001/pos-web/shipment/create', data =json.dumps(param), headers = headers, name = "API下单", catch_response = True ) as response:
            if response.status_code ==201:
                response.success()



import six
from itertools import chain

from flask import request, Response
from locust import stats as locust_stats, runners as locust_runners
from locust import User, task, events
from prometheus_client import Metric, REGISTRY, exposition

# This locustfile adds an external web endpoint to the locust master, and makes it serve as a prometheus exporter.
# Runs it as a normal locustfile, then points prometheus to it.
# locust -f prometheus_exporter.py --master

# Lots of code taken from [mbolek's locust_exporter](https://github.com/mbolek/locust_exporter), thx mbolek!


class LocustCollector(object):
    registry = REGISTRY

    def __init__(self, environment, runner):
        self.environment = environment
        self.runner = runner

    def collect(self):
        # collect metrics only when locust runner is spawning or running.
        runner = self.runner

        if runner and runner.state in (locust_runners.STATE_SPAWNING, locust_runners.STATE_RUNNING):
            stats = []
            for s in chain(locust_stats.sort_stats(runner.stats.entries), [runner.stats.total]):
                stats.append({
                    "method": s.method,
                    "name": s.name,
                    "num_requests": s.num_requests,
                    "num_failures": s.num_failures,
                    "avg_response_time": s.avg_response_time,
                    "min_response_time": s.min_response_time or 0,
                    "max_response_time": s.max_response_time,
                    "current_rps": s.current_rps,
                    "median_response_time": s.median_response_time,
                    "ninetieth_response_time": s.get_response_time_percentile(0.9),
                    # only total stats can use current_response_time, so sad.
                    #"current_response_time_percentile_95": s.get_current_response_time_percentile(0.95),
                    "avg_content_length": s.avg_content_length,
                    "current_fail_per_sec": s.current_fail_per_sec
                })

            # perhaps StatsError.parse_error in e.to_dict only works in python slave, take notices!
            errors = [e.to_dict() for e in six.itervalues(runner.stats.errors)]

            metric = Metric('locust_user_count', 'Swarmed users', 'gauge')
            metric.add_sample('locust_user_count', value=runner.user_count, labels={})
            yield metric

            metric = Metric('locust_errors', 'Locust requests errors', 'gauge')
            for err in errors:
                metric.add_sample('locust_errors', value=err['occurrences'],
                                  labels={'path': err['name'], 'method': err['method'],
                                          'error': err['error']})
            yield metric

            is_distributed = isinstance(runner, locust_runners.MasterRunner)
            if is_distributed:
                metric = Metric('locust_slave_count', 'Locust number of slaves', 'gauge')
                metric.add_sample('locust_slave_count', value=len(runner.clients.values()), labels={})
                yield metric

            metric = Metric('locust_fail_ratio', 'Locust failure ratio', 'gauge')
            metric.add_sample('locust_fail_ratio', value=runner.stats.total.fail_ratio, labels={})
            yield metric

            metric = Metric('locust_state', 'State of the locust swarm', 'gauge')
            metric.add_sample('locust_state', value=1, labels={'state': runner.state})
            yield metric

            stats_metrics = ['avg_content_length', 'avg_response_time', 'current_rps', 'current_fail_per_sec',
                             'max_response_time', 'ninetieth_response_time', 'median_response_time', 'min_response_time',
                             'num_failures', 'num_requests']

            for mtr in stats_metrics:
                mtype = 'gauge'
                if mtr in ['num_requests', 'num_failures']:
                    mtype = 'counter'
                metric = Metric('locust_stats_' + mtr, 'Locust stats ' + mtr, mtype)
                for stat in stats:
                    # Aggregated stat's method label is None, so name it as Aggregated
                    # locust has changed name Total to Aggregated since 0.12.1
                    if 'Aggregated' != stat['name']:
                        metric.add_sample('locust_stats_' + mtr, value=stat[mtr],
                                          labels={'path': stat['name'], 'method': stat['method']})
                    else:
                        metric.add_sample('locust_stats_' + mtr, value=stat[mtr],
                                          labels={'path': stat['name'], 'method': 'Aggregated'})
                yield metric


@events.init.add_listener
def locust_init(environment, runner, **kwargs):
    print("locust init event received")
    if environment.web_ui and runner:
        @environment.web_ui.app.route("/export/prometheus")
        def prometheus_exporter():
            registry = REGISTRY
            encoder, content_type = exposition.choose_encoder(request.headers.get('Accept'))
            if 'name[]' in request.args:
                registry = REGISTRY.restricted_registry(request.args.get('name[]'))
            body = encoder(registry)
            return Response(body, content_type=content_type)
        REGISTRY.register(LocustCollector(environment, runner))


class websitUser(HttpUser):
    tasks = [Test]
    host = "http://120.79.131.69:20000"#Kparcel生产环境测试地址
    # host = "http://testtmss.vaiwan.com/hello/"
    min_wait = 1000  # 单位为毫秒
    max_wait = 2000  # 单位为毫秒
