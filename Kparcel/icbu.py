import json
import random

import openpyxl
import requests




Kparcel = "http://120.79.229.241:8000"
def open_excel():#读取ececl数据
    wb = openpyxl.load_workbook('icbu_draft_orders.xlsx')#TH泰国参数
    ws = wb.active
    data = []
    for i in range(4,5):#ws.max_row
        x=str(ws['A'+str(i)].value)
        data.append(json.loads(x))
    return data

def Icbu_Create():
    datas = open_excel()
    url = Kparcel+"/tms-saas-web/order/booking"
    #url = "http://172.16.3.155:8998/web/order/booking"
    for i in range(len(datas)):
        data1 = datas[i]
        data1['bookingOrderDTO']['aliOrderNo']="ALS145634"+str(random.randint(1000000,99999999))
        print(data1['bookingOrderDTO']['aliOrderNo'])
        payload=json.dumps(data1)
        headers = {
          'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        print(response.text)
    return data1['bookingOrderDTO']['aliOrderNo']

Icbu_Create()